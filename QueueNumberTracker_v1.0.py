import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from PIL import Image, ImageTk
import pyautogui
import cv2
import numpy as np
import yaml
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import webbrowser
import re
import os
import sys
import threading
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO
import easyocr
import torch
from matplotlib.dates import DateFormatter, MinuteLocator
import matplotlib.ticker as ticker

# 最後更新時間： 2026/3/14

# 中英版
texts = {
    "en": {
        "screenCaptureControlGui": "Screen Capture Control GUI",
        "successConnection": "Successful Connection",
        "writeRecognitionResult": "Write Recognition Result",
        "timeStamp": "Timestamp",
        "errorOccurred": "Error",
        "serialNumber": "Queue Number",
        "time": "Time",
        "interval": "Interval",
        "fileCreatedAt": "File Created At",
        "fileCreationFailed": "File Creation Failed",
        "correctedVersion": "Corrected Version",
        "validNumber": "Valid Number",
        "recordCount": "Record Count",
        "validCount": "Valid Count",
        "filteredDataCount": "Filtered Data Count",
        "average": "Average",
        "standardDeviation": "Standard Deviation",
        "coefficientOfVariation": "Coefficient of Variation",
        "recognitionResult": "Recognition Result",
        "dataSource": "Data Source",
        "error": "Error",
        "executionTimeReachedStopRecognition": ">>> Execution Time Reached, Stop Recognition",
        "selectHospital": "Hospital",
        "selectDoctor": "Doctor",
        "hospitalUrl": "Hospital URL",
        "processedImage": "Processed Image",
        "ok": "OK",
        "createFile": "Create File",
        "runDuration": "Run Duration",
        "recognitionInterval": "Recog. Interval",
        "return": "Return",
        "systemInProgress": "System In Progress",
        "abort": "Abort",
        "imageRecognitionCompleted": "Image Recognition Completed",
        "dataProcessingThreshold": "Data Processing Threshold",
        "threshold1": "Threshold 1",
        "threshold2": "Threshold 2",
        "isGenerateImage": "Generate Image",
        "startProcessing": "Start",
        "exitProgram": "Exit",
        "Finish":"Corrected data has been generated",
        "Confirm":"confirm",
        "adaptive": "Adaptive",
        "manual": "Manual",
        "threshold_label": "Threshold Method:",
        "file_name_empty": "File name cannot be empty",
        "file_name_invalid_chars": "File name contains invalid characters: \\ / : * ? \" < > |",
        "file_exists": "File '{file_name}.xlsx' already exists, please choose another name",
        "invalid_timer_range": "Execution time must be an integer between 1 and 10 hours",
        "invalid_interval_range": "Recognition interval must be an integer between 1 and 600 seconds",
        "Picture_1_Title":"Consultation Time vs. Queue Number",
        "consultation_time":"Consultation Time",
        "Picture_2_Title":"Consultation Duration vs. Queue Number",
        "consultation_duration":"Consultation Duration (minutes)",
        "Picture_3_Title":"Valid Number vs. Valid Interval",
        "valid_interval":"Valid Interval (minutes)",
        "Picture_4_Title":"Consultation Duration vs. Valid Interval",
    },

    "zh": {
        "screenCaptureControlGui": "螢幕擷取控制 GUI",
        "successConnection": "成功連接",
        "writeRecognitionResult": "寫入辨識結果",
        "timeStamp": "時間戳記",
        "errorOccurred": "發生錯誤",
        "serialNumber": "號次",
        "time": "時間",
        "interval": "間隔",
        "fileCreatedAt": "已建立檔案於",
        "fileCreationFailed": "建立檔案失敗",
        "correctedVersion": "校正版",
        "validNumber": "有效編號",
        "recordCount": "資料數量",
        "validCount": "有效數量",
        "filteredDataCount": "過濾數量",
        "average": "平均",
        "standardDeviation": "標準差",
        "coefficientOfVariation": "變異係數",
        "recognitionResult": "辨識結果",
        "dataSource": "資料出處",
        "error": "錯誤",
        "executionTimeReachedStopRecognition": ">>> 已達設定執行時間，結束辨識",
        "selectHospital": "選擇醫院",
        "selectDoctor": "選擇醫生",
        "hospitalUrl": "醫院網址",
        "processedImage": "處理後影像",
        "ok": "確定",
        "createFile": "建立檔案",
        "runDuration": "執行時間",
        "recognitionInterval": "辨識間隔",
        "return": "返回",
        "systemInProgress": "系統運作中",
        "abort": "中止執行",
        "imageRecognitionCompleted": "影像辨識結束",
        "dataProcessingThreshold": "設定數據處裡閥值參數",
        "threshold1": "(排除網頁刷新)閥值一",
        "threshold2": "(排除過號號次)閥值二",
        "isGenerateImage": "產生圖片",
        "startProcessing": "開始處理",
        "exitProgram": "結束程式",
        "Finish":"已生成校正後數據",
        "Confirm":"確認資料",
        "adaptive": "自適應",
        "manual": "手動設定",
        "threshold_label": "二值化方式",
        "file_name_empty": "檔名不能為空",
        "file_name_invalid_chars": "檔名包含不允許的字元: \\ / : * ? \" < > |",
        "file_exists": "檔案 '{file_name}.xlsx' 已存在，請更換檔名",
        "invalid_timer_range": "執行時間必須為 1~10 小時之間的正整數",
        "invalid_interval_range": "辨識間隔必須為 1~600 秒之間的正整數",
        "Picture_1_Title":"看診時間與號次",
        "consultation_time":"看診時間",
        "Picture_2_Title":"看診時數與號次",
        "consultation_duration":"看診時數(分鐘)",
        "Picture_3_Title":"有效編號與有效間隔",
        "valid_interval":"有效間隔(分鐘)",
        "Picture_4_Title":"看診時數與有效間隔",
    }
}

# 取得在各種執行環境下可用的資源檔路徑
def get_resource_path(filename: str) -> str:
    # 取得目前程式所在目錄（兼容 .py 與 exe）
    if getattr(sys, 'frozen', False):  
        # base_path = sys._MEIPASS # 非 onefile 編譯模式路徑
        base_path = os.path.dirname(sys.executable)
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))  
    return os.path.join(base_path, filename)

# 取得 YAML 檔案完整路徑
print("Connecting to YAML files...")
yaml_path = get_resource_path("hospital_doctors.yaml")

# 載入 YAML 資料
try:
    with open(yaml_path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
except FileNotFoundError:
    messagebox.showerror("Error", f"YAML file not found: {yaml_path}")
    sys.exit(1)
except Exception as e:
    messagebox.showerror("Error", f"An unexpected error occurred while reading YAML: {e}")
    sys.exit(1)

# 控制語言變數: "en" or "zh", 預設英文
lang = data.get("language", "en")
if  lang not in  ["en", "zh"]:
    print("Language setting error. Only English or Traditional Chinese is supported. It will be automatically switched to English.")
    print(lang)
    lang = "en"

# 取醫院資料（過濾掉 "language" key）
hospital_data = {k: v for k, v in data.items() if k != "language" and k!= "lastBBox"}

# 將來的 Excel 先在全域宣告
file_path = None  

# 建立主視窗
root = tk.Tk()
root.title(texts[lang]["screenCaptureControlGui"])
root.geometry("720x480")

stop_event = threading.Event()  # 全局停止旗標

# 第三頁關閉視窗
def on_closing():
    stop_event.set()
    # 判斷是按鈕關閉還是視窗關閉可以用參數或拆成兩個函式
    # 這裡假設是視窗關閉事件，直接關閉視窗
    root.destroy()

# 第三頁停止執行 => go_to_page4()
def on_close_button():
    stop_event.set()
    go_to_page4()

# 攔截視窗關閉事件
root.protocol("WM_DELETE_WINDOW", on_closing)

''' -------------------- 涵式區 -------------------- '''
# 函數：在文字框中新增輸出
def append_output(text):
    text_output.config(state="normal")      # 解除唯讀
    text_output.insert("end", text + "\n")  # 插入新文字並加上換行
    text_output.see("end")                  # 自動捲到最底
    text_output.config(state="disabled")    # 保持唯讀狀態

def capture_screen(region):
    # region:擷取螢幕範圍
    screenshot = pyautogui.screenshot(region=region)
    image = cv2.cvtColor(np.array(screenshot), cv2.COLOR_RGB2BGR)
    
    return image

# 2025/10/20 EasyOCR
# 辨識範圍:設定的語言+數字
def create_reader(lang='en'):
    # 檢查 GPU 是否可用
    gpu_available = torch.cuda.is_available()

    # 印出結果
    print(f"CUDA GPU available: {gpu_available}")

    try:
        if lang == 'zh':
            reader = easyocr.Reader(['ch_tra', 'en'], gpu=gpu_available)
        else:
            reader = easyocr.Reader([lang], gpu=gpu_available)

    except Exception as e:
        print(f"Error during EasyOCR initialization: {e}")
        print("Falling back to English model ('en')...")

        # 確保即使失敗也能回傳一個最基礎的 Reader
        try:
            return easyocr.Reader(['en'], gpu=False)
        except Exception as e3:
            print(f"Critical Error: Unable to initialize OCR engine. {e3}")
            return None
        
    return reader

print("Loading OCR models. This may take a moment on first run...")
reader = create_reader(lang)
if reader is None:
    error_msg = (
        "Fatal Error: Unable to initialize the OCR engine.\n\n"
        "Possible reasons:\n"
        "1. Corrupted model files\n"
        "2. Insufficient memory (RAM)\n"
        "3. No internet connection for first run\n\n"
        "The application will now close.")
    messagebox.showerror("Critical Error", error_msg)
    print("Failed to initialize OCR. Exiting...")
    sys.exit(1)  

print("OCR models loaded successfully!")

def recognize_digits(image, return_image=0):
    # 2025/11/25更新,使用 Easy OCR 不用二值化
    # 限制只辨識數字
    results = reader.readtext(image, detail=0, allowlist='0123456789')

    # 整合成一串數字
    result = ''.join(results)

    if return_image:
        return image, result
    else:
        return result

def write_to_excel(latest_number, now_time, file_path):
    """
    將兩個參數寫入 Excel 檔案的最下一行。
        latest_number (int): 要寫入的數字。
        now_time (str): 當下時間，格式可以是字符串。
        file_path (str): Excel 檔案的路徑。
    """
    try:
        # 嘗試加載現有 Excel 文件
        workbook = load_workbook(file_path)
        sheet = workbook.active  # 使用默認的第一個工作表
        print(f'成功連接：{file_path}')
        append_output(f'{texts[lang]["successConnection"]}:{file_path}')

        latest_number = latest_number.strip()   # 刪除字串頭尾的換行和空白字元
        sheet.append([latest_number, now_time]) # 找到最後一行，在其後新增一行
        
        # 保存 Excel 文件
        workbook.save(file_path)
        print(f"寫入 {file_path}: {latest_number}, {now_time}") 
        append_output(f'{texts[lang]["writeRecognitionResult"]}: {latest_number}, {texts[lang]["timeStamp"]}: {now_time}') 

    except Exception as e:
        # 捕捉其他所有例外，並印出錯誤訊息
        print(f'發生錯誤: {e}')
        append_output(f'{texts[lang]["errorOccurred"]}: {e}')
        return 1    # 觸發主程式 break

    return 0

# 檢查目前是否打包成 exe 並切換路徑
def get_save_folder():
    if getattr(sys, 'frozen', False):  # 判斷是否為打包模式
        base_path = os.path.dirname(sys.executable)  # exe 所在資料夾
    else:
        base_path = os.path.abspath(".")
    
    save_folder = os.path.join(base_path, "data")

    # 如果 data 資料夾不存在，就建立它
    if not os.path.exists(save_folder):
        os.makedirs(save_folder)

    return save_folder

# 更新第一頁顯示畫面
def update_capture():
    # 擷取螢幕範圍（使用 pyautogui 擷取）
    x, y = capture_params['x'], capture_params['y']
    w, h = capture_params['width'], capture_params['height']
    frame = capture_screen((x, y, w, h))
    frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)  # 轉回 Tkinter 可用格式

    ''' 將 frame 縮放成固定大小 (選取範圍時不要壓縮圖片比較好)
    frame = cv2.resize(frame, (DISPLAY_WIDTH, DISPLAY_HEIGHT))  
    '''

    img = Image.fromarray(frame)
    imgtk = ImageTk.PhotoImage(image=img)

    page1_video_label.imgtk = imgtk
    page1_video_label.configure(image=imgtk)

    # 每 100ms 更新一次
    page1_video_label.after(100, update_capture)

# 刷新醫生選單(包含網址)
def update_doctor_menu(event=None):
    hospital_info = hospital_data.get(selected_hospital.get(), {})
    doctors = hospital_info.get("doctors", [])
    url = hospital_info.get("url", "None")
    
    doctor_menu["values"] = doctors
    if doctors:
        selected_doctor.set(doctors[0])
    else:
        selected_doctor.set("None")

    hospital_url_var.set(f"{url}")

# URL 超連結
def open_url(event):
    url = hospital_url_var.get()
    if url:
        webbrowser.open(url)

# 滑鼠移入：改變樣式（加底線、改顏色）
def on_enter(event):
    url_label_2.config(font=("Arial", 12, "underline"), fg="darkblue")

# 滑鼠移出：恢復原來樣式
def on_leave(event):
    url_label_2.config(font=("Arial", 12), fg="blue")
 
# 滑桿變更事件
def slider_changed(param, val):
    capture_params[param] = int(val)
    entry_vars[param].set(str(val))  # 同步 Entry 顯示

# 檢查第一頁圖大小是否大於第二頁顯示空間，如果有，等比例壓縮
def Check_image_size(width, height, pil_img):
    if width > 360 or height > 240:
        scale_w = 360 / width
        scale_h = 240 / height
        scale = min(scale_w, scale_h)  # 取較小比例，確保不超過限制
        new_width = int(width * scale)
        new_height = int(height * scale)
        pil_img = pil_img.resize((new_width, new_height), resample=Image.Resampling.LANCZOS)
    
    return pil_img

# 建立 Excel，回傳 file_path
def build_Execl(timer_val, file_name):
    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    h = selected_hospital.get()
    d = selected_doctor.get()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Dr.{d}"           # 工作表名稱
    sheet.append([date, f"{timer_val} h", f"{h}"]) # 日期、時長、醫院
    sheet.append([f'{texts[lang]["serialNumber"]}',
                  f'{texts[lang]["time"]}',
                  f'{texts[lang]["interval"]}']) 

    # 對 A1:C1 置中
    for row in sheet["A1:C2"]:  # 兩列都置中
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 保存 Excel 文件
    try:
        # 將 Excel 存在資料夾 data
        save_folder = get_save_folder()
        file_path = os.path.join(save_folder, f"{file_name}.xlsx")

        workbook.save(file_path)
        print(f"已建立檔案於: {file_path}")
        append_output(f'{texts[lang]["fileCreatedAt"]}: {file_path}')

        return file_path

    except Exception as e:
        print(f"建立檔案失敗: \n{e}")
        append_output(f'{texts[lang]["fileCreationFailed"]}: \n{e}')

# 資料處裡演算法
def Data_processing_algorithms(file_path):
    # 使用第一個分頁作為來源
    wb = load_workbook(file_path)
    source_ws = wb.worksheets[0]

    # 從第 3 列開始加公式(第一頁)
    for row in range(3, source_ws.max_row):
        formula = f"=(B{row+1}-B{row})*86400"
        source_ws.cell(row=row, column=3).value = formula

    # 建立新的工作表，若已存在名為「校正版」的工作表就刪掉（避免重複）
    if texts[lang]["correctedVersion"] in wb.sheetnames:
        del wb[texts[lang]["correctedVersion"]]
    target_ws = wb.create_sheet(texts[lang]["correctedVersion"])

    # 先複製標頭（前兩列）
    for row_index in [1, 2]:
        row_data = [source_ws.cell(row=row_index, column=col).value for col in range(1, 8)]
        target_ws.append(row_data)
    
    # 校正資料
    max_row = source_ws.max_row
    last_number = -1            # 上一筆號次
    DPM = 0                     # 偏移量
    threshold_of_refresh = int(Threshold_1_input.get()) # 刷新錯誤時間閥值
    threshold_of_Pass = int(Threshold_2_input.get())    # 過號時間閥值

    ''' 校正演算法 + 生成校正資料 '''
    for row in range(3, max_row + 1):
        row_data = []  # 每列要寫入的資料

        # 號次
        try:
            value = int(source_ws.cell(row=row, column=1).value)
        except:
            print(f"第 {row} 列異常：號次 = {source_ws.cell(row=row, column=1).value}，視為雜訊忽略不計")
            DPM += 1
            continue

        # 號次異常條件檢查
        if  value < 0 or value > 200 or value == last_number:
            DPM += 1
            continue

        row_data.append(value)

        # 時間
        try:
            time_value = source_ws.cell(row=row, column=2).value
            next_time_value = source_ws.cell(row=row+1, column=2).value
            time_1 = datetime.strptime(time_value, "%H:%M:%S")
            time_2 = datetime.strptime(next_time_value, "%H:%M:%S")

        except:
            if  row == max_row:
                time_1 = time_2
            else:
                print("時間數據欄位出現格式錯誤")

        time_diff = int((time_2 - time_1).total_seconds())    # 計算時間差

        # 檢查是否小於等於 threshold_of_refresh 秒，(最後一筆資料時間差=0)
        if 0 < time_diff and time_diff <= threshold_of_refresh:
            print(f"第 {row} 列異常：時間差 {time_diff} 秒，視為雜訊忽略不計")
            DPM += 1
            continue
        
        # (演算邏輯)確定沒有刷新錯誤後才更新參數:"上一筆數字"
        last_number = value
        row_data.append(time_value)

        # 間隔（公式，注意要修正偏移）
        formula = f"=(B{(row - DPM) + 1}-B{(row - DPM)})*86400"
        row_data.append(formula)

        # 以 row 為單位加入
        target_ws.append(row_data)

    ''' <<< 統計數據 >>> '''

    # 儲存檔案以讀取計算後數據
    wb.save(file_path)
    wb = load_workbook(file_path)
    target_ws = wb[texts[lang]["correctedVersion"]]
    max_row = target_ws.max_row

    valid_data = 1 # 有效數據編號
    chart2_y_data = []

    for row in range(4, max_row):
            
        time_value = target_ws.cell(row=row, column=2).value
        next_time_value = target_ws.cell(row=row+1, column=2).value
        time_1 = datetime.strptime(time_value, "%H:%M:%S")
        time_2 = datetime.strptime(next_time_value, "%H:%M:%S")
        time_diff = int((time_2 - time_1).total_seconds())    # 計算時間差

        # 檢查是否小於等於 threshold_of_refresh 秒，(最後一筆資料時間差=0)
        if time_diff > threshold_of_Pass:
            target_ws[f'D{row}'] = valid_data
            valid_data += 1
            chart2_y_data.append(time_diff)

    target_ws['D2'] = texts[lang]["validNumber"]
    target_ws['F1'] = texts[lang]["recordCount"]
    target_ws['G1'] = texts[lang]["validCount"]
    target_ws['H1'] = texts[lang]["filteredDataCount"]

    target_ws['F2'] = f'=COUNTA(A3:A{max_row})'                              # 資料數量
    target_ws['G2'] = f'=COUNTIFS(C4:C{max_row-1}, ">={threshold_of_Pass}")' # 有效資料數量
    target_ws['H2'] = '=F2-G2'

    target_ws['F5'] = texts[lang]["average"]
    target_ws['G5'] = texts[lang]["standardDeviation"]
    target_ws['H5'] = texts[lang]["coefficientOfVariation"]

    target_ws['F6'] = f'=AVERAGE(C4:C{max_row-1})'  # 平均值
    target_ws['G6'] = f'=STDEV(C4:C{max_row-1})'    # 標準差
    target_ws['H6'] = '=F6/F2'                      # 變異系數

    # 設定置中對齊
    center_alignment = Alignment(horizontal='center', vertical='center')

    for row in target_ws.iter_rows():
        for cell in row:
            if cell.value is not None:  # 只對有資料的儲存格套用對齊
                cell.alignment = center_alignment

    # 提示語
    if lang == "zh":
        target_ws['F4'] = f"過濾 頭尾 & 過號 (間隔小於 {threshold_of_Pass} s) 的數據所計算的結果" 
        target_ws.append([])
        target_ws.append([f'以上為程式自動校正，過濾不正常 & 間隔小於 {threshold_of_refresh} s 的號次'])

    else:
        target_ws['F4'] = f"Results calculated after filtering head/tail passes and duplicate passes (interval < {threshold_of_Pass} s)"
        target_ws.append([])
        target_ws.append([f"The above results were automatically corrected by the program, filtering abnormal records and intervals < {threshold_of_refresh} s"])
    
    user_choice = image_option.get()
    ''' <<< 使用 matplotlib 繪圖 >>> '''
    if user_choice == "yes":

        plt.rcParams['font.family'] = 'Microsoft JhengHei'  # 設定中文字型
        plt.rcParams['axes.unicode_minus'] = False          # 避免負號顯示錯誤（常見中文+數字混合問題）

        ''' 圖片一：看診時間與號次'''
        # 擷取資料（x:號碼, y:時間）
        x_data, y_data= [], []

        for row in target_ws.iter_rows(min_row=3, max_row=max_row, min_col=1, max_col=2):
            x_val = row[1].value
            y_val = row[0].value
            if x_val and y_val:
                # str時間轉時間
                if isinstance(x_val, str):
                    x_val = datetime.strptime(x_val, "%H:%M:%S")        
                x_data.append(x_val)
                y_data.append(y_val)

        plt.figure(figsize=(10, 6))
        plt.plot(x_data, y_data, marker='o', linestyle='-')
        plt.title(texts[lang]["Picture_1_Title"], fontsize=20)
        plt.xlabel(texts[lang]["consultation_time"], fontsize=16)
        plt.ylabel(texts[lang]["serialNumber"], fontsize=16)
        plt.grid(True)

        # 固定時間間距設定 
        ax = plt.gca()
        # 每 10 分鐘 一個刻度
        ax.xaxis.set_major_locator(MinuteLocator(interval=10))
        # 時間格式：顯示成 HH:mm
        ax.xaxis.set_major_formatter(DateFormatter('%H:%M'))

        # 儲存成圖片(在 memory)
        buf1 = BytesIO()
        plt.savefig(buf1, format='png', bbox_inches='tight')
        plt.close()

        # 插入圖片到 Excel
        buf1.seek(0)    # 將 檔案指標 移到開頭
        img1 = ExcelImage(buf1)
        img1.anchor = "H10"
        target_ws.add_image(img1)

        ''' 圖片二：看診時數與號次'''
        # 擷取資料（x:號次, y:時間）
        x_data, y_data= [], []

        for row in target_ws.iter_rows(min_row=3, max_row=max_row, min_col=1, max_col=2):
            time_str = row[1].value   # 時間
            number = row[0].value     # 號次

            if time_str and number:
                # 解析時間
                t = datetime.strptime(time_str, "%H:%M:%S")
                x_data.append(t)
                y_data.append(number)

        # 轉成從 0 分鐘開始的軸
        first_time = x_data[0]

        # 把 datetime → 分鐘差（float）
        x_minutes = [(t - first_time).total_seconds() / 60 for t in x_data]

        plt.figure(figsize=(10, 6))
        plt.plot(x_minutes, y_data, marker='o', linestyle='-')
        plt.title(texts[lang]["Picture_2_Title"], fontsize=20)
        plt.xlabel(texts[lang]["consultation_duration"], fontsize=16)
        plt.ylabel(texts[lang]["serialNumber"], fontsize=16)
        plt.grid(True)

        # 每 10 分鐘刻度
        ax = plt.gca()
        ax.xaxis.set_major_locator(ticker.MultipleLocator(10))  # 每 10 分鐘刻度

        buf2 = BytesIO()
        plt.savefig(buf2, format='png', bbox_inches='tight')
        plt.close()
        buf2.seek(0)
        img2 = ExcelImage(buf2)
        img2.anchor = "V10"
        target_ws.add_image(img2)

        ''' 圖片三：有效編號與間隔'''
        # 擷取資料（x:號碼, y:時間）
        x_data = []
        chart2_y_data = [i / 60 for i in chart2_y_data] # 將有效間隔化為分鐘單位

        # 擷取資料：X 軸為 D 欄，Y 軸為 C 欄，排除 D 欄為空
        for row in target_ws.iter_rows(min_row=3, max_row=max_row, min_col=3, max_col=4):
            x_val = row[1].value  
            if x_val is not None:
                x_data.append(x_val)

        plt.figure(figsize=(10, 6))
        plt.plot(x_data, chart2_y_data, marker='o', linestyle='-')
        plt.title(texts[lang]["Picture_3_Title"], fontsize=20)
        plt.xlabel(texts[lang]["validNumber"], fontsize=16)
        plt.ylabel(texts[lang]["valid_interval"], fontsize=16)
        plt.grid(True)

        ay = plt.gca()
        ay.yaxis.set_major_locator(ticker.MultipleLocator(2))  # 每 2 分鐘刻度

        buf3 = BytesIO()
        plt.savefig(buf3, format='png', bbox_inches='tight')
        plt.close()
        buf3.seek(0)
        img3 = ExcelImage(buf3)
        img3.anchor = "H40"
        target_ws.add_image(img3)

        ''' 圖片四：看診時間與有效間隔'''
        # 擷取資料（x:號碼, y:時間）
        x_data = []

        # 擷取資料：X 軸為排除 D 欄為空的 B 欄，Y 軸為 D 欄
        for row in target_ws.iter_rows(min_row=3, max_row=max_row, min_col=2, max_col=4):

            x_val = row[0].value
            if x_val and row[2].value is not None:
                t = datetime.strptime(x_val, "%H:%M:%S")
                x_data.append(t)
                

        # 轉成從 0 分鐘開始的軸
        first_time = x_data[0]

        # 把 datetime → 分鐘差（float）
        x_minutes = [(t - first_time).total_seconds() / 60 for t in x_data]

        plt.figure(figsize=(10, 6))
        plt.plot(x_minutes, chart2_y_data, marker='o', linestyle='-')
        plt.title(texts[lang]["Picture_4_Title"], fontsize=20)
        plt.xlabel(texts[lang]["consultation_duration"], fontsize=16)
        plt.ylabel(texts[lang]["valid_interval"], fontsize=16)
        plt.grid(True)

        axy = plt.gca()
        axy.xaxis.set_major_locator(ticker.MultipleLocator(10))  # 每 10 分鐘刻度
        axy.yaxis.set_major_locator(ticker.MultipleLocator(2))  # 每 2 分鐘刻度

        buf4 = BytesIO()
        plt.savefig(buf4, format='png', bbox_inches='tight')
        plt.close()
        buf4.seek(0)
        img4 = ExcelImage(buf4)
        img4.anchor = "V40"
        target_ws.add_image(img4)

    # 儲存檔案
    wb.save(file_path)
    if lang == "zh":
        print(f"已完成檔案 {file_path} 校正。")
    else:
        print(f"File {file_path} has been calibrated.")

# 2025/10/13 ver0.8 更新:刷新左側圖片、辨識結果
def update_page2_image():
    # 從 capture_params 讀取參數
    x = capture_params['x']
    y = capture_params['y']
    width = capture_params['width']
    height = capture_params['height']

    region = (x, y, width, height)
    captured_image = capture_screen(region)

    # 根據目前設定執行辨識（包含自適應或手動閾值）
    image, digits = recognize_digits(captured_image, 1)

    # OpenCV(BGR) → PIL(RGB)
    rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

    # 轉成 Tkinter 可顯示格式
    pil_img = Image.fromarray(rgb_image)
    pil_img = Check_image_size(width, height, pil_img)
    photo_img = ImageTk.PhotoImage(pil_img)

    binary_img_canvas.delete("all")
    binary_img_canvas.create_image(180, 120, image=photo_img, anchor="center")
    binary_img_canvas.image = photo_img

    # 辨識結果垂直滾動條
    digit_result_text.delete("1.0", tk.END)    # 刷新前需要先清空     
    digit_result_text.insert(tk.END, f"{texts[lang]['recognitionResult']}:{digits}")

''' --------------------< 頁面跳轉函式 >-------------------- '''
def go_to_page1():
    page2_frame.pack_forget()                   # pack 暫時不顯示，設定還在
    page1_frame.pack(fill="both", expand=True)  # 填滿畫面、容器空間變大時跟著撐開

def go_to_page2():
    # 從 capture_params 讀取參數
    x = capture_params['x']
    y = capture_params['y']
    width = capture_params['width']
    height = capture_params['height']

    # 更新 YAML 內容中的 lastBBox
    data['lastBBox'] = {
        'X': x,
        'Y': y,
        'W': width,
        'H': height
    }

    # 寫回 yaml
    # 打包後 exe 版本需寫到可寫路徑（例如使用者資料夾）
    if getattr(sys, 'frozen', False):
        save_path = os.path.join(os.path.expanduser("~"), "hospital_doctors.yaml")
    else:
        save_path = get_resource_path("hospital_doctors.yaml")

    with open(save_path, "w", encoding="utf-8") as f:
        yaml.safe_dump(data, f, allow_unicode=True, sort_keys=False)

    # 刷新左側圖片、辨識結果
    update_page2_image()

    h = selected_hospital.get()
    d = selected_doctor.get()

    if lang == "zh":
        page2_check_label.config(text=f"{texts[lang]['dataSource']}\n{h}\n{d} 醫師")
    elif lang == "en":
        page2_check_label.config(text=f"{texts[lang]['dataSource']}\n{h}\n Dr.{d}")
    
    # 計算檔案名稱（日期 + 部門）
    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    department = h.split("_")[1]
    file_name = f"{date}_{department}_{d}"
    file_name_var.set(file_name)    # 輸入框預設值

    page1_frame.pack_forget()
    page2_frame.pack(fill="both", expand=True)

def go_to_page3():
    global file_path

    # 取得第二頁輸入值確保其合法
    file_name = file_name_var.get()
    timer_val = Timer_input.get()
    interval_val = interval_input.get()

    # 檔名檢查
    if not file_name.strip():
        messagebox.showerror(texts[lang]["error"], texts[lang]["file_name_empty"])
        return

    # 不允許的字元 (Windows 常見不合法字元)
    invalid_chars = r'[\\/:*?"<>|]'
    if re.search(invalid_chars, file_name):
        messagebox.showerror(texts[lang]["error"], texts[lang]["file_name_invalid_chars"])
        return
    
    # 檢查檔案是否已存在
    save_folder = get_save_folder()  
    file_full_path = os.path.join(save_folder, f"{file_name}.xlsx")  
    if os.path.exists(file_full_path):
        messagebox.showerror(texts[lang]["error"], texts[lang]["file_exists"].format(file_name=file_name))
        return
    
    # 檢查是否為正整數且符合範圍
    if not timer_val.isdigit() or not (0 < int(timer_val) <= 10):
        messagebox.showerror(texts[lang]["error"], texts[lang]["invalid_timer_range"])
        Timer_entry.focus_set()
        return  

    if not interval_val.isdigit() or not (0 < int(interval_val) <= 600):
        messagebox.showerror(texts[lang]["error"], texts[lang]["invalid_interval_range"])
        interval_entry.focus_set()
        return  
    
    # 建立 Excel   
    file_path = build_Execl(timer_val, file_name)

    page2_frame.pack_forget()
    page3_frame.pack(fill="both", expand=True)

    ''' --------------------< 以下辨識迴圈(副執行序) >-------------------- '''
    def recognition_loop():
        timer_value = int(Timer_input.get())
        interval_value = int(interval_input.get())
        total_seconds = timer_value * 3600
        start_time = time.time()
        region = (capture_params['x'], capture_params['y'], capture_params['width'], capture_params['height'])
        latest_number = -1

        while  not stop_event.is_set():
            elapsed = time.time() - start_time
            if elapsed > total_seconds:
                append_output(texts[lang]["executionTimeReachedStopRecognition"])
                break

            captured_image = capture_screen(region)
            digits = recognize_digits(captured_image)

            print(f"辨識結果: {digits}")
            append_output(f"{texts[lang]['recognitionResult']}: {digits}")

            if digits != latest_number:
                latest_number = digits
                now = datetime.now()
                now_time = now.strftime("%H:%M:%S")
                if write_to_excel(latest_number, now_time, file_path):
                    break

            # 分段睡眠，方便快速響應關閉事件
            sleep_time = 0
            while sleep_time < interval_value:
                if stop_event.is_set():
                    break
                time.sleep(0.2)
                sleep_time += 0.2

    # 加開執行緒給辨識迴圈，以免 sleep 導致 GUI 卡死
    thread = threading.Thread(target=recognition_loop, daemon=True)
    thread.start()  

    ''' --------------------< 以上辨識迴圈 >-------------------- '''

def go_to_page4():
    page3_frame.pack_forget()
    page4_frame.pack(fill="both", expand=True)

def go_to_page5():
    try:
        Data_processing_algorithms(file_path)
    except Exception as e:
        print(f"發生錯誤：{str(e)}")
        label5_text.set(f"{texts[lang]['errorOccurred']}：\n{str(e)}")

    page4_frame.pack_forget()
    page5_frame.pack(fill="both", expand=True)

''' ---------------------------------------- 以下 GUI ---------------------------------------- '''

''' 
--------------------<<< 第一頁 >>>-------------------- 
'''

page1_frame = tk.Frame(root)
page1_frame.pack(fill="both", expand=True)

''' ----------< 以下選單 >---------- '''
hospital_names = list(hospital_data.keys())

# 建立一個子 frame 來放在同一列的元件
row_frame = tk.Frame(page1_frame)
row_frame.pack(pady=(20,0))

selected_hospital = tk.StringVar()          # 醫院選單變數
selected_hospital.set(hospital_names[0])    # 預設第一個
selected_doctor = tk.StringVar()            # 醫生選單變數
hospital_url_var = tk.StringVar()           # 網址變數

# 醫院 Label、選單
hospital_label = tk.Label(row_frame, text=f"{texts[lang]['selectHospital']}:", font=("Arial", 18))
hospital_label.grid(row=0, column=0, padx=5)
hospital_menu = ttk.Combobox(row_frame, textvariable=selected_hospital, values=hospital_names, state="readonly", font=("Arial", 18), width=21)
hospital_menu.grid(row=0, column=1, padx=10)

# 醫生 Label、選單
doctor_label = tk.Label(row_frame, text=f"{texts[lang]['selectDoctor']}:", font=("Arial", 18))
doctor_label.grid(row=0, column=2, padx=5)
doctor_menu = ttk.Combobox(row_frame, textvariable=selected_doctor, state="readonly", font=("Arial", 18), width=7)
doctor_menu.grid(row=0, column=3, padx=10)

# 網址顯示 Label（新增）醫院網址: 
url_label_1 = tk.Label(row_frame, text=f"{texts[lang]['hospitalUrl']}:", font=("Arial", 18))
url_label_1.grid(row=1, column=0, padx=5)
url_label_2 = tk.Label(row_frame, textvariable=hospital_url_var, font=("Arial", 12), fg="blue", anchor="w", justify="left", cursor="hand2", wraplength=570)
url_label_2.grid(row=1, column=1, columnspan=4, sticky="w", pady=(0, 0))

# 超連結按鈕(綁定滑鼠事件)
url_label_2.bind("<Button-1>", open_url)    # 點擊
url_label_2.bind("<Enter>", on_enter)       # 滑入
url_label_2.bind("<Leave>", on_leave)       # 滑出

# 選擇醫院時重新呼叫 update_doctor_menu
hospital_menu.bind("<<ComboboxSelected>>", update_doctor_menu)

# 初始化醫生選單
update_doctor_menu()

''' ----------< 以上選單 >----------'''

# 左側顯示畫面
page1_video_label = tk.Label(page1_frame)
page1_video_label.pack(side=tk.LEFT, padx=(20,0))

# 右側控制區
control_frame = tk.Frame(page1_frame)
control_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(0,20))

''' ----------< 以下拉桿 >----------'''

# 建立四個拉桿與標籤
sliders = {}
value_labels = {}
slider_range = {
    'x': (0, 1920),
    'y': (0, 1080),
    'width': (200, 480),
    'height': (200, 360)
}

# 拉桿範圍參數
capture_params = {
    'x': 0,
    'y': 0,
    'width': 480,
    'height': 360
}

# 如果 YAML 有紀錄上次圈選範圍，就覆蓋預設值
if 'lastBBox' in data:
    bbox = data['lastBBox']
    capture_params['x'] = bbox.get('X', capture_params['x'])
    capture_params['y'] = bbox.get('Y', capture_params['y'])
    capture_params['width'] = bbox.get('W', capture_params['width'])
    capture_params['height'] = bbox.get('H', capture_params['height'])

sliders = {}
entry_vars = {}  # 儲存文字輸入框的變數

for idx, (param, (min_val, max_val)) in enumerate(slider_range.items()):
    row_base = idx * 2  # 每個參數占兩行：一行 Entry，一行滑桿

    # Label設定
    tk.Label(control_frame, text=f"{param}：", width=7, anchor="center").grid(
        row=row_base, column=0, sticky="w", padx=5, pady=(20, 7))

    # 輸入框設定
    var = tk.StringVar()
    var.set(str(capture_params[param]))
    entry = tk.Entry(control_frame, textvariable=var, width=8, justify="center")
    entry.grid(row=row_base, column=1, sticky="w", pady=(20, 7))

    # 滑桿設定
    slider = tk.Scale(
        control_frame,
        from_=min_val,
        to=max_val,
        orient=tk.HORIZONTAL,
        command=lambda val, p=param: slider_changed(p, val),
        showvalue=False,  # 隱藏上方數值
        length=150
    )

    slider.set(capture_params[param])
    slider.grid(row=row_base + 1, column=0, columnspan=2, sticky="we", padx=5)

    # Entry 改變時同步滑桿
    def entry_callback(event, p=param, v=var, s=slider):
        try:
            new_val = int(v.get())
            new_val = max(min(new_val, slider_range[p][1]), slider_range[p][0])
            capture_params[p] = new_val
            s.set(new_val)
        except ValueError:
            v.set(str(capture_params[p]))  # 無效輸入時還原

    entry.bind("<Return>", entry_callback)
    entry.bind("<FocusOut>", entry_callback)

    # 儲存元件
    sliders[param] = slider
    entry_vars[param] = var

''' ----------< 以上拉桿 >---------- '''

# 確定按鈕
btn_confirm = tk.Button(control_frame,text=texts[lang]["ok"], command=go_to_page2, font=("Arial", 18, "bold"), padx=10, pady=5, relief="raised", bg="#F44336", fg="white", activebackground="#D32F2F", bd=5)
btn_confirm.place(relx=1.0, rely=1.0, anchor="se", x=0, y=-10)

'''
--------------------<<< 第二頁 >>>--------------------
'''
page2_frame = tk.Frame(root)

# 上方顯示畫面
page2_up_frame = tk.Frame(page2_frame)
page2_up_frame.pack(pady=(15,0))

# 左側顯示畫面
page2_left_frame = tk.Frame(page2_frame)
page2_left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(20,0))
page2_left_up = tk.Frame(page2_left_frame)
page2_left_up.pack(pady=10, anchor="nw")
page2_left_down = tk.Frame(page2_left_frame)
page2_left_down.pack(fill="both", pady=(0,10))

# 右側控制區
page2_right_frame = tk.Frame(page2_frame)
page2_right_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(0,20))
page2_right_up = tk.Frame(page2_right_frame)
page2_right_up.pack(pady=10, anchor="ne")    # 靠右上排列

''' ----------< 以下 第二頁 上方 >---------- '''
# 檔名輸入框
file_name_var = tk.StringVar(value="file_name")
tk.Label(page2_up_frame, text=f"{texts[lang]['createFile']} :", font=("Arial", 20)).pack(side="left", padx=(0, 10))
tk.Entry(page2_up_frame, textvariable=file_name_var, font=("Arial", 18), width=30).pack(side="left")
tk.Label(page2_up_frame, text=".xlsx", font=("Arial", 20)).pack(side="left", padx=(10, 0))

''' ----------< 以下 第二頁 左側 >---------- '''
image_text_label = tk.Label(page2_left_up, text=texts[lang]["processedImage"], font=("Arial", 20))
image_text_label.pack(anchor="center")

# 處理後畫面
binary_img_canvas = tk.Canvas(page2_left_up, width=360, height=240, bg="white", highlightthickness=0)
binary_img_canvas.pack(side=tk.TOP)

# 辨識結果(滾動條)
digit_result_text = tk.Text(page2_left_down, font=("Arial", 18), height=1, width=23, wrap="word")
scrollbar = tk.Scrollbar(page2_left_down, command=digit_result_text.yview)
digit_result_text.configure(yscrollcommand=scrollbar.set)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
digit_result_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

''' ----------< 以下 第二頁 右側 >---------- '''
# 顯示選擇的醫院與醫生
page2_check_label = tk.Label(page2_right_up, text="", font=("Arial", 20)) # text 更新於涵式：update_page2_content()
page2_check_label.pack(pady=(10,0))

# === 二值化設定區 === 2025/10/13 ver 0.8 更新
threshold_frame = tk.Frame(page2_right_frame)
threshold_frame.pack(anchor="center", pady=(10, 10))
'''
tk.Label(threshold_frame, text=texts[lang]["threshold_label"], font=("Arial", 20)).pack(anchor="w")

# 換行：選項與輸入框
option_frame = tk.Frame(threshold_frame)
option_frame.pack(anchor="w", pady=(5, 0))

use_adaptive = tk.BooleanVar(value=True)   # 預設使用自適應
manual_threshold = tk.StringVar(value="")  # 儲存手動輸入值

# 自適應選項（單選）
adaptive_radio = tk.Radiobutton(
    option_frame,
    text=texts[lang]["adaptive"],
    variable=use_adaptive,
    value=True,
    font=("Arial", 20)
)
adaptive_radio.pack(side="left", padx=(5, 10))

# 手動閾值選項（單選 + 輸入框）
manual_radio = tk.Radiobutton(
    option_frame,
    text=texts[lang]["manual"],
    variable=use_adaptive,
    value=False,
    font=("Arial", 20)
)
manual_radio.pack(side="left")

manual_entry = tk.Entry(threshold_frame, textvariable=manual_threshold, width=5, font=("Arial", 13))
manual_entry.pack(side="left", padx=(5, 0))
tk.Label(threshold_frame, text="(0~255)", font=("Arial", 15)).pack(side="left", padx=(5, 0))

# 當選項改變時重新刷新影像
use_adaptive.trace_add("write", lambda *args: update_page2_image())

# 當焦點離開輸入框時觸發更新
manual_entry.bind("<FocusOut>", lambda e: update_page2_image())
'''

# 計時器輸入框
Timer_frame = tk.Frame(page2_right_frame)
Timer_frame.pack(padx=(5, 10))
tk.Label(Timer_frame, text=f"{texts[lang]['runDuration']}：", font=("Arial", 20)).pack(side="left")
Timer_input = tk.StringVar(value="3")    # 執行時間預設3小時
Timer_entry = tk.Entry(Timer_frame, textvariable=Timer_input, font=("Arial", 20), width=3, justify="right")
Timer_entry.pack(side="left", padx=(5, 10))
tk.Label(Timer_frame, text="h", font=("Arial", 20)).pack(side="left")

# 間隔輸入框
interval_frame = tk.Frame(page2_right_frame)
interval_frame.pack(padx=(5, 10))
tk.Label(interval_frame, text=f"{texts[lang]['recognitionInterval']}：", font=("Arial", 20)).pack(side="left")
interval_input = tk.StringVar(value="10")    # 間隔時間預設10s
interval_entry = tk.Entry(interval_frame, textvariable=interval_input, font=("Arial", 20), width=3, justify="right")
interval_entry.pack(side="left", padx=(5, 10))
tk.Label(interval_frame, text="s", font=("Arial", 20)).pack(side="left")

# 開始執行、返回按鈕
btn_start = tk.Button(page2_frame, text=texts[lang]["ok"], command=go_to_page3, font=("Arial", 18, "bold"), padx=10, pady=5, relief="raised", bg="#F44336", fg="white", activebackground="#D32F2F", bd=5)
btn_start.place(relx=1.0, rely=1.0, anchor="se", x=-20, y=-10)
btn_back = tk.Button(page2_frame, text=texts[lang]["return"], command=go_to_page1, font=("Arial", 18, "bold"), padx=10, pady=5, relief="raised", bg="#9E9E9E", fg="white", activebackground="#7E7E7E", bd=5)
btn_back.place(relx=0.0, rely=1.0, anchor="sw", x=20, y=-10)

'''
--------------------<<< 第三頁 >>>--------------------
新增第三頁輸出時，使用涵式 append_output(value), type:'str'
'''
page3_frame = tk.Frame(root)

label3 = tk.Label(page3_frame, text=texts[lang]["systemInProgress"], font=("Arial", 24))
label3.pack(pady=(20,10))

''' ----------< 以下輸出面板 >----------'''
'''
規格：
    wrap="word":               換行時以「單字」為單位，而不是硬切字元。
    bg="black", fg="white":    文字框背景黑色、字體白色，模仿終端機樣式。
    insertbackground="white":  設定游標顏色為白色（輸入模式時可見）。
    font=("Consolas", 12):     採用等寬字體 Consolas,12 號字，方便對齊。
    state="disabled":          預設禁用輸入，避免使用者手動改內容。
    pack(side="left", fill="both", expand=True)：放在左邊並填滿空間
'''

output_frame = tk.Frame(page3_frame, bg="black", bd=2, relief="sunken")
output_frame.pack(padx=40, pady=(0,80), fill="both", expand=True)

text_output = tk.Text(output_frame, wrap="word", bg="black", fg="white", insertbackground="white", font=("Consolas", 12), state="disabled")
text_output.pack(side="left", fill="both", expand=True)

# 滾動條
scrollbar = tk.Scrollbar(output_frame, command=text_output.yview)
scrollbar.pack(side="right", fill="y")
text_output.config(yscrollcommand=scrollbar.set)

''' ----------< 以上輸出面板 >----------'''

# 按鈕
btn_stop_page3 = tk.Button(page3_frame, text=texts[lang]["abort"], command=on_close_button, font=("Arial", 18, "bold"), padx=10, pady=5, relief="groove", bg="#F44336", fg="white", activebackground="#D32F2F", bd=7)
btn_stop_page3.place(relx=0.5, rely=1.0, anchor="s", y=-10)

'''
--------------------<<< 第四頁 >>>--------------------
'''
page4_frame = tk.Frame(root)

label4 = tk.Label(page4_frame, text=texts[lang]["imageRecognitionCompleted"], font=("Arial", 32))
label4.pack(fill="x", pady=(20,10))

# 顯示數據處理設定
settings_label = tk.Label(page4_frame, text=texts[lang]["dataProcessingThreshold"], font=("Arial", 24), anchor="w")
settings_label.pack(fill="x", padx=50, pady=(10,0))

# 閥值一(網頁刷新錯誤) 輸入框
Threshold_1_frame = tk.Frame(page4_frame)
Threshold_1_frame.pack(pady=10)
tk.Label(Threshold_1_frame, text=texts[lang]["threshold1"], font=("Arial", 24)).pack(side="left")
Threshold_1_input = tk.StringVar(value="15")    # 預設 15 s
Threshold_1_entry = tk.Entry(Threshold_1_frame, textvariable=Threshold_1_input, font=("Arial", 24), width=3, justify='center')
Threshold_1_entry.pack(side="left", padx=(5, 10))
tk.Label(Threshold_1_frame, text="s", font=("Arial", 24)).pack(side="left")

# 閥值二輸入框
Threshold_2_frame = tk.Frame(page4_frame)
Threshold_2_frame.pack()
tk.Label(Threshold_2_frame, text=texts[lang]["threshold2"], font=("Arial", 24)).pack(side="left")
Threshold_2_input = tk.StringVar(value="65")    # 預設 65 s
Threshold_2_entry = tk.Entry(Threshold_2_frame, textvariable=Threshold_2_input, font=("Arial", 24), width=3, justify='center')
Threshold_2_entry.pack(side="left", padx=(5, 10))
tk.Label(Threshold_2_frame, text="s", font=("Arial", 20)).pack(side="left")

# 生成圖片? (選項)
option_label = tk.Label(page4_frame, text=texts[lang]["isGenerateImage"], font=("Arial", 24))
option_label.pack(pady=(20, 5))

image_option = tk.StringVar(value="yes")  # 預設 yes
option_frame = tk.Frame(page4_frame)
option_frame.pack()

yes_btn = tk.Radiobutton(option_frame, text="Yes", variable=image_option, value="yes", font=("Arial", 20))
yes_btn.pack(side="left", padx=5)
no_btn = tk.Radiobutton(option_frame, text="No", variable=image_option, value="no", font=("Arial", 20))
no_btn.pack(side="left", padx=5)

# 按鈕 
btn_start_page4 = tk.Button(page4_frame, text=texts[lang]["startProcessing"], command=go_to_page5, font=("Arial", 18, "bold"), padx=10, pady=5, relief="raised", bg="#F44336", fg="white", activebackground="#D32F2F", bd=7)
btn_start_page4.place(relx=1.0, rely=1.0, anchor="se", x=-150, y=-10)
btn_stop_page4 = tk.Button(page4_frame, text=texts[lang]["exitProgram"], command=root.destroy, font=("Arial", 18, "bold"), padx=10, pady=5, relief="raised", bg="#9E9E9E", fg="white", activebackground="#D32F2F", bd=7)
btn_stop_page4.place(relx=0.0, rely=1.0, anchor="sw", x=150, y=-10)
'''
--------------------<<< 第五頁 >>>--------------------
'''
page5_frame = tk.Frame(root)

label5_text = tk.StringVar()
label5_text.set(texts[lang]["Finish"]) 

label5 = tk.Label(page5_frame, textvariable=label5_text, font=("Arial", 32), wraplength=600, justify="left")
label5.pack(pady=150)

btn_stop_page5 = tk.Button(page5_frame, text=texts[lang]["exitProgram"], command=root.destroy, font=("Arial", 18, "bold"), padx=10, pady=5, relief="groove", bg="#F44336", fg="white", activebackground="#D32F2F", bd=7)
btn_stop_page5.place(relx=0.5, rely=1.0, anchor="s", y=-10)

# 開始畫面更新
update_capture()

root.mainloop()